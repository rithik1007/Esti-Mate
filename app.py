from flask import Flask, render_template, request, jsonify
import os
from datetime import datetime
import json
import requests
from requests.auth import HTTPBasicAuth
import base64
from dotenv import load_dotenv

# Import AI modules
from ai_modules.design_generator.solution_designer import SolutionDesigner, DesignApprovalWorkflow
from ai_modules.code_generator.code_generator import AICodeGenerator
from ai_modules.repo_analyzer.codebase_analyzer import CodebaseAnalyzer
from ai_modules.ai_estimator import AIEstimator

load_dotenv()

app = Flask(__name__)

# Configuration
app.config['SECRET_KEY'] = 'your-secret-key-here'

class JiraClient:
    def __init__(self):
        self.base_url = None
        self.token = None
        self.email = None
    
    def configure(self, base_url, email, token):
        self.base_url = base_url.rstrip('/')
        self.email = email
        self.token = token
    
    def get_ticket_details(self, ticket_key):
        if not all([self.base_url, self.email, self.token]):
            raise Exception("JIRA configuration missing. Please configure JIRA settings first.")
        
        url = f"{self.base_url}/rest/api/2/issue/{ticket_key}?expand=changelog"
        auth = HTTPBasicAuth(self.email, self.token)
        
        print(f"Fetching JIRA ticket from: {url}")
        print(f"Using email: {self.email}")
        
        try:
            response = requests.get(url, auth=auth, timeout=10)
        except requests.exceptions.Timeout:
            raise Exception("JIRA request timed out. Please check your JIRA server URL and try again.")
        except requests.exceptions.ConnectionError:
            raise Exception("Cannot connect to JIRA server. Please check your JIRA base URL and internet connection.")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Network error while connecting to JIRA: {str(e)}")
        
        if response.status_code == 404:
            raise Exception(f"Ticket '{ticket_key}' not found. Please verify the ticket number exists and you have permission to view it.")
        elif response.status_code == 401:
            raise Exception("Authentication failed. Please check your JIRA email and API token are correct.")
        elif response.status_code == 403:
            raise Exception(f"Access denied to ticket '{ticket_key}'. You don't have permission to view this ticket.")
        elif response.status_code == 400:
            raise Exception(f"Invalid ticket format '{ticket_key}'. Please use format like 'PROJ-123'.")
        elif response.status_code != 200:
            raise Exception(f"JIRA server error ({response.status_code}). Please try again later or contact your JIRA administrator.")
        
        data = response.json()
        
        # Extract comments
        comments = []
        if 'comment' in data['fields'] and data['fields']['comment']['comments']:
            for comment in data['fields']['comment']['comments']:
                comments.append({
                    'author': comment['author']['displayName'],
                    'body': comment['body'],
                    'created': comment['created']
                })
        
        # Extract labels
        labels = data['fields'].get('labels', [])
        
        # Extract linked issues
        linked_issues = []
        if 'issuelinks' in data['fields']:
            for link in data['fields']['issuelinks']:
                if 'outwardIssue' in link:
                    linked_issues.append({
                        'key': link['outwardIssue']['key'],
                        'summary': link['outwardIssue']['fields']['summary'],
                        'type': link['type']['outward']
                    })
                if 'inwardIssue' in link:
                    linked_issues.append({
                        'key': link['inwardIssue']['key'],
                        'summary': link['inwardIssue']['fields']['summary'],
                        'type': link['type']['inward']
                    })
        
        # Extract fix versions
        fix_versions = []
        if 'fixVersions' in data['fields']:
            for version in data['fields']['fixVersions']:
                fix_versions.append({
                    'name': version['name'],
                    'released': version.get('released', False)
                })
        
        # Extract changelog for status transitions
        status_history = []
        time_in_status = {}
        if 'changelog' in data:
            for history in data['changelog']['histories']:
                for item in history['items']:
                    if item['field'] == 'status':
                        status_history.append({
                            'from': item.get('fromString', ''),
                            'to': item.get('toString', ''),
                            'changed_at': history['created'],
                            'author': history['author']['displayName']
                        })
        
        # Calculate time spent in each status
        if status_history:
            from datetime import datetime
            from dateutil import parser
            
            for i, transition in enumerate(status_history):
                status_name = transition['from']
                try:
                    start_time = parser.parse(transition['changed_at'])
                    
                    if i + 1 < len(status_history):
                        end_time = parser.parse(status_history[i + 1]['changed_at'])
                    else:
                        end_time = datetime.now(start_time.tzinfo)
                    
                    duration_hours = (end_time - start_time).total_seconds() / 3600
                    
                    if status_name not in time_in_status:
                        time_in_status[status_name] = 0
                    time_in_status[status_name] += duration_hours
                except Exception as e:
                    print(f"DEBUG - Error parsing date {transition['changed_at']}: {e}")
                    continue
        
        # Extract time tracking data
        time_tracking = {}
        if 'timetracking' in data['fields']:
            time_tracking = {
                'original_estimate': data['fields']['timetracking'].get('originalEstimate', ''),
                'remaining_estimate': data['fields']['timetracking'].get('remainingEstimate', ''),
                'time_spent': data['fields']['timetracking'].get('timeSpent', ''),
                'original_estimate_seconds': data['fields']['timetracking'].get('originalEstimateSeconds', 0),
                'time_spent_seconds': data['fields']['timetracking'].get('timeSpentSeconds', 0)
            }
        
        return {
            'key': data['key'],
            'summary': data['fields']['summary'],
            'description': data['fields'].get('description', ''),
            'issue_type': data['fields']['issuetype']['name'],
            'priority': data['fields'].get('priority', {}).get('name', 'Medium'),
            'status': data['fields']['status']['name'],
            'comments': comments,
            'labels': labels,
            'linked_issues': linked_issues,
            'fix_versions': fix_versions,
            'status_history': status_history,
            'time_in_status': time_in_status,
            'time_tracking': time_tracking,
            'created': data['fields'].get('created', ''),
            'updated': data['fields'].get('updated', '')
        }

jira_client = JiraClient()
# Auto-configure from environment variables
if all([os.getenv('JIRA_BASE_URL'), os.getenv('JIRA_EMAIL'), os.getenv('JIRA_API_TOKEN')]):
    jira_client.configure(
        os.getenv('JIRA_BASE_URL'),
        os.getenv('JIRA_EMAIL'),
        os.getenv('JIRA_API_TOKEN')
    )

class ProjectEstimator:
    def __init__(self):
        self.phase_weights = {
            'requirements': 0.15,
            'design': 0.20,
            'development': 0.48,
            'testing': 0.15,
            'deployment': 0.02
        }
    
    def estimate_project(self, description, jira_number=None, jira_data=None):
        """Estimate project phases based on description and JIRA data"""
        
        # Analyze complexity with JIRA data
        complexity = self._analyze_complexity(description, jira_data)
        base_hours = self._get_base_hours(complexity, jira_data)
        
        # Calculate phase estimates
        estimates = {}
        for phase, weight in self.phase_weights.items():
            estimates[phase] = round(base_hours * weight, 1)
        
        result = {
            'jira_number': jira_number,
            'description': description,
            'complexity': complexity,
            'total_hours': base_hours,
            'phases': estimates,
            'estimated_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        if jira_data:
            result['jira_details'] = {
                'issue_type': jira_data.get('issue_type'),
                'priority': jira_data.get('priority', 'Medium'),
                'status': jira_data.get('status')
            }
        
        return result
    
    def _analyze_complexity(self, description, jira_data=None):
        """Analyze project complexity based on keywords and JIRA data"""
        description_lower = description.lower()
        complexity_score = 0
        
        # Original keyword analysis
        high_complexity_keywords = [
            'integration', 'api', 'database', 'migration', 'security',
            'authentication', 'microservice', 'complex', 'multiple systems'
        ]
        
        medium_complexity_keywords = [
            'crud', 'form', 'validation', 'report', 'dashboard',
            'ui', 'frontend', 'backend'
        ]
        
        high_count = sum(1 for keyword in high_complexity_keywords if keyword in description_lower)
        medium_count = sum(1 for keyword in medium_complexity_keywords if keyword in description_lower)
        
        complexity_score += high_count * 2 + medium_count
        
        # JIRA-based complexity factors
        if jira_data:
            issue_type = jira_data.get('issue_type', '').lower()
            priority = jira_data.get('priority', 'medium').lower()
            
            # Issue type complexity
            if issue_type in ['epic', 'story']:
                complexity_score += 2
            elif issue_type in ['task', 'improvement']:
                complexity_score += 1
            elif issue_type == 'bug':
                complexity_score += 0.5
            
            # Priority impact
            if priority in ['critical', 'highest']:
                complexity_score += 1.5
            elif priority in ['high', 'major']:
                complexity_score += 1
        
        # Description length factor
        if len(description.split()) > 50:
            complexity_score += 1
        
        # Determine final complexity
        if complexity_score >= 4:
            return 'High'
        elif complexity_score >= 2:
            return 'Medium'
        else:
            return 'Low'
    
    def _get_base_hours(self, complexity, jira_data=None):
        """Get base hours based on complexity and JIRA factors"""
        complexity_hours = {
            'Low': 40,
            'Medium': 80,
            'High': 160
        }
        
        base_hours = complexity_hours.get(complexity, 80)
        
        # Adjust based on JIRA factors
        if jira_data:
            issue_type = jira_data.get('issue_type', '').lower()
            priority = jira_data.get('priority', 'medium').lower()
            
            # Issue type multiplier
            if issue_type == 'epic':
                base_hours *= 1.5
            elif issue_type == 'bug':
                base_hours *= 0.7
            
            # Priority multiplier
            if priority in ['critical', 'highest']:
                base_hours *= 1.2
            elif priority in ['low', 'lowest']:
                base_hours *= 0.8
        
        return round(base_hours)

estimator = ProjectEstimator()

# Initialize AI components
design_generator = SolutionDesigner(
    api_key=os.getenv('AZURE_OPENAI_API_KEY'),
    azure_endpoint=os.getenv('AZURE_OPENAI_ENDPOINT')
)
code_generator = AICodeGenerator(
    api_key=os.getenv('AZURE_OPENAI_API_KEY'),
    azure_endpoint=os.getenv('AZURE_OPENAI_ENDPOINT')
)
codebase_analyzer = CodebaseAnalyzer()
approval_workflow = DesignApprovalWorkflow()
# Determine AI provider
ai_provider = os.getenv('AI_PROVIDER', 'azure_openai')

# Prepare AWS config for Amazon Q
aws_config = {
    'region': os.getenv('AWS_REGION'),
    'access_key': os.getenv('AWS_ACCESS_KEY_ID'),
    'secret_key': os.getenv('AWS_SECRET_ACCESS_KEY'),
    'app_id': os.getenv('AMAZON_Q_APP_ID')
}

ai_estimator = AIEstimator(
    api_key=os.getenv('AZURE_OPENAI_API_KEY'),
    azure_endpoint=os.getenv('AZURE_OPENAI_ENDPOINT'),
    ai_provider=ai_provider,
    aws_config=aws_config if ai_provider == 'amazon_q' else None
)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/bulk-estimate-page')
def bulk_estimate_page():
    return render_template('bulk_estimate.html')

@app.route('/ai-workflow')
def ai_workflow():
    return render_template('ai_workflow.html')

@app.route('/configure-jira', methods=['POST'])
def configure_jira():
    try:
        data = request.get_json()
        base_url = data.get('base_url')
        email = data.get('email')
        token = data.get('token')
        
        if not all([base_url, email, token]):
            return jsonify({'error': 'All JIRA fields are required'}), 400
        
        jira_client.configure(base_url, email, token)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/estimate', methods=['POST'])
def estimate():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data received'}), 400
            
        description = data.get('description', '')
        jira_number = data.get('jira_number', '').strip()
        
        # Auto-detect: if JIRA number is provided, try to fetch from JIRA
        use_jira = bool(jira_number)
        
        jira_data = None
        if jira_number:
            try:
                # Try to fetch from JIRA
                jira_data = jira_client.get_ticket_details(jira_number)
                description = f"{jira_data['summary']}. {jira_data['description']}"
                jira_number = jira_data['key']
            except Exception as e:
                # If JIRA fetch fails but description is provided, continue with description
                if description:
                    print(f"DEBUG - JIRA fetch failed, using provided description: {e}")
                    jira_data = None
                else:
                    # If no description and JIRA fails, return error
                    return jsonify({
                        'error': str(e),
                        'error_type': 'jira_error'
                    }), 400
        
        if not description:
            return jsonify({'error': 'Description is required'}), 400
        
        # Check if AI estimation is requested
        use_ai = data.get('use_ai', False)
        selected_phases = data.get('selected_phases', {
            'requirements': True,
            'design': True,
            'development': True,
            'testing': True,
            'deployment': True
        })
        
        phase_percentages = data.get('phase_percentages', {
            'requirements': 15,
            'design': 20,
            'development': 48,
            'testing': 15,
            'deployment': 2
        })
        
        custom_phases = data.get('custom_phases', {})
        actual_hours = data.get('actual_hours')  # For learning system
        uses_ai_tools = data.get('uses_ai_tools', False)  # AI tools checkbox
        
        print(f"\n========== ESTIMATION REQUEST DEBUG ==========")
        print(f"DEBUG - use_ai from request: {use_ai} (type: {type(use_ai)})")
        print(f"DEBUG - Azure API key exists: {bool(os.getenv('AZURE_OPENAI_API_KEY'))}")
        print(f"DEBUG - Azure API key length: {len(os.getenv('AZURE_OPENAI_API_KEY', ''))}")
        print(f"DEBUG - Azure endpoint: {os.getenv('AZURE_OPENAI_ENDPOINT')}")
        print(f"DEBUG - ai_estimator.api_key exists: {bool(ai_estimator.api_key)}")
        print(f"DEBUG - ai_estimator.client exists: {bool(ai_estimator.client)}")
        print(f"DEBUG - Condition check: use_ai={use_ai} AND api_key={bool(os.getenv('AZURE_OPENAI_API_KEY'))}")
        print(f"DEBUG - Will use AI: {use_ai and os.getenv('AZURE_OPENAI_API_KEY')}")
        print(f"============================================\n")
        
        if use_ai and os.getenv('AZURE_OPENAI_API_KEY'):
            print("DEBUG - Using AI estimation with FAB model")
            # Add AI tools flag to JIRA data for processing
            if jira_data:
                jira_data['uses_ai_tools'] = uses_ai_tools
            else:
                jira_data = {'uses_ai_tools': uses_ai_tools}
            
            # Use AI estimation
            ai_result = ai_estimator.estimate_with_ai(description, jira_data)
            
            # Check if status filtering should override custom percentages
            status = jira_data.get('status', '').lower() if jira_data else ''
            status_override = status in ['qa', 'sit', 'testing', 'ready for testing', 'in testing', 'uat', 'user acceptance testing', 'ready for deployment', 'staging', 'done', 'closed', 'resolved', 'deployed', 'production', 'in progress', 'development', 'coding', 'in development']
            
            # Check if user has customized phases - if so, don't apply status filtering
            has_custom_phases = bool(custom_phases) or any(phase_percentages.get(phase, 0) != default_percent for phase, default_percent in [('requirements', 15), ('design', 20), ('development', 48), ('testing', 15), ('deployment', 2)])
            
            if status_override and not has_custom_phases:
                print(f"DEBUG - Status '{status}' detected, applying status-based filtering (no custom phases)")
                # Apply status filtering to AI result first
                ai_result = ai_estimator._apply_status_based_filtering(ai_result, jira_data)
                filtered_phases = ai_result.get('phases', {})
                total_filtered_hours = ai_result.get('total_hours', 0)
            else:
                if has_custom_phases:
                    print(f"DEBUG - Custom phases detected, skipping status-based filtering")
                # Apply custom percentages and filter phases
                base_total = ai_result.get('total_hours', 80)
                print(f"DEBUG - AI base total hours: {base_total}")
                print(f"DEBUG - Phase percentages: {phase_percentages}")
                
                filtered_phases = {}
                total_filtered_hours = 0
                
                # Process standard phases
                for phase in ['requirements', 'design', 'development', 'testing', 'deployment']:
                    if selected_phases.get(phase, True):
                        custom_hours = round(base_total * (phase_percentages.get(phase, 0) / 100), 2)
                        print(f"DEBUG - {phase}: {phase_percentages.get(phase, 0)}% of {base_total} = {custom_hours} hours")
                        filtered_phases[phase] = custom_hours
                        total_filtered_hours += custom_hours
                
                # Process custom phases
                for phase_key in custom_phases.keys():
                    if selected_phases.get(phase_key, True):
                        custom_hours = round(base_total * (phase_percentages.get(phase_key, 0) / 100), 2)
                        print(f"DEBUG - {phase_key}: {phase_percentages.get(phase_key, 0)}% of {base_total} = {custom_hours} hours")
                        filtered_phases[phase_key] = custom_hours
                        total_filtered_hours += custom_hours
                
                # Round total to avoid floating point precision issues
                total_filtered_hours = round(total_filtered_hours, 2)
                print(f"DEBUG - Total filtered hours: {total_filtered_hours}")
            
            # Get historical analysis if JIRA data available
            historical_analysis = None
            if jira_data and jira_data.get('status_history'):
                historical_analysis = ai_estimator._analyze_jira_historical_patterns(jira_data)
            
            # Format result to match existing structure
            estimate_result = {
                'jira_number': jira_number,
                'description': description,
                'complexity': ai_result.get('complexity', 'Medium'),
                'total_hours': total_filtered_hours,
                'phases': filtered_phases,
                'estimated_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'estimation_method': ai_result.get('estimation_method', 'ai_powered'),
                'ai_confidence': ai_result.get('confidence', 75),
                'ai_reasoning': ai_result.get('reasoning', ''),
                'risk_factors': ai_result.get('risk_factors', []),
                'custom_phase_names': custom_phases,
                'historical_analysis': historical_analysis,
                'testing_breakdown': ai_result.get('testing_breakdown', {})
            }
            
            if jira_data:
                estimate_result['jira_details'] = {
                    'issue_type': jira_data.get('issue_type'),
                    'priority': jira_data.get('priority', 'Medium'),
                    'status': jira_data.get('status'),
                    'status_history': jira_data.get('status_history', []),
                    'time_in_status': jira_data.get('time_in_status', {})
                }
        else:
            print("\n========== USING RULE-BASED ESTIMATION ==========")
            print(f"DEBUG - Reason: use_ai={use_ai}, api_key_exists={bool(os.getenv('AZURE_OPENAI_API_KEY'))}")
            print(f"================================================\n")
            # Use rule-based estimation
            estimate_result = estimator.estimate_project(description, jira_number, jira_data)
            
            # Apply custom percentages and filter phases
            base_total = estimate_result['total_hours']
            filtered_phases = {}
            total_filtered_hours = 0
            
            # Process standard phases
            for phase in ['requirements', 'design', 'development', 'testing', 'deployment']:
                if selected_phases.get(phase, True):
                    custom_hours = round(base_total * (phase_percentages.get(phase, 0) / 100), 1)
                    filtered_phases[phase] = custom_hours
                    total_filtered_hours += custom_hours
            
            # Process custom phases
            for phase_key in custom_phases.keys():
                if selected_phases.get(phase_key, True):
                    custom_hours = round(base_total * (phase_percentages.get(phase_key, 0) / 100), 1)
                    filtered_phases[phase_key] = custom_hours
                    total_filtered_hours += custom_hours
            
            estimate_result['phases'] = filtered_phases
            estimate_result['total_hours'] = round(total_filtered_hours, 2)
            estimate_result['estimation_method'] = 'rule_based'
            estimate_result['custom_phase_names'] = custom_phases
        
        # Record estimation for learning
        if jira_number:
            ai_estimator.learning_system.record_estimation(
                jira_ticket=jira_number,
                estimated_hours=estimate_result['total_hours'],
                complexity=estimate_result['complexity'],
                phases=estimate_result['phases'],
                method=estimate_result.get('estimation_method', 'unknown'),
                description=description,
                actual_hours=actual_hours
            )
        
        return jsonify(estimate_result)
    except Exception as e:
        print(f"Error in estimate: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/history')
def history():
    return jsonify([])

@app.route('/generate-design', methods=['POST'])
def generate_design():
    try:
        data = request.get_json()
        ticket_key = data.get('ticket_key')
        repo_path = data.get('repo_path', '')
        
        if not ticket_key:
            return jsonify({'error': 'Ticket key is required'}), 400
        
        jira_data = jira_client.get_ticket_details(ticket_key)
        
        codebase_context = {}
        if repo_path and os.path.exists(repo_path):
            codebase_context = codebase_analyzer.analyze_codebase(repo_path)
        
        design = design_generator.generate_solution_design(jira_data, codebase_context)
        
        approvers = data.get('approvers', ['tech_lead', 'architect'])
        approval_id = approval_workflow.submit_for_approval(ticket_key, design, approvers)
        
        return jsonify({
            'design': design,
            'approval_id': approval_id,
            'status': 'pending_approval',
            'ticket_key': ticket_key
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/approve-design', methods=['POST'])
def approve_design():
    try:
        data = request.get_json()
        approval_id = data.get('approval_id')
        approver = data.get('approver')
        approved = data.get('approved', False)
        comment = data.get('comment', '')
        
        if not all([approval_id, approver]):
            return jsonify({'error': 'Approval ID and approver are required'}), 400
        
        approval_workflow.add_approval_comment(approval_id, approver, comment, approved)
        
        if approved:
            success = approval_workflow.approve_design(approval_id, approver)
            if success:
                return jsonify({'status': 'approved', 'message': 'Design approved successfully'})
            else:
                return jsonify({'error': 'Failed to approve design'}), 400
        else:
            return jsonify({'status': 'feedback_provided', 'message': 'Feedback added to design'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate-code', methods=['POST'])
def generate_code():
    try:
        data = request.get_json()
        approval_id = data.get('approval_id')
        repo_path = data.get('repo_path', '')
        
        if not approval_id:
            return jsonify({'error': 'Approval ID is required'}), 400
        
        approved_design_data = approval_workflow.get_approved_design(approval_id)
        if not approved_design_data:
            return jsonify({'error': 'Design not found or not approved'}), 400
        
        design = approved_design_data['design']
        ticket_key = approved_design_data['ticket_key']
        
        jira_data = jira_client.get_ticket_details(ticket_key)
        
        if repo_path and os.path.exists(repo_path):
            generated_code = code_generator.generate_code_from_design(design, repo_path, jira_data)
        else:
            generated_code = code_generator.generate_code_from_design(design, '', jira_data)
        
        return jsonify({
            'generated_code': generated_code,
            'ticket_key': ticket_key,
            'design_id': approval_id,
            'status': 'code_generated'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/analyze-codebase', methods=['POST'])
def analyze_codebase():
    try:
        data = request.get_json()
        repo_path = data.get('repo_path')
        
        if not repo_path:
            return jsonify({'error': 'Repository path is required'}), 400
        
        if not os.path.exists(repo_path):
            return jsonify({'error': 'Repository path does not exist'}), 400
        
        analysis = codebase_analyzer.analyze_codebase(repo_path)
        
        return jsonify({
            'analysis': analysis,
            'repo_path': repo_path,
            'status': 'analysis_complete'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/pending-designs')
def get_pending_designs():
    try:
        pending = approval_workflow.get_pending_designs()
        return jsonify({'pending_designs': pending})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/learning-dashboard')
def learning_dashboard():
    return render_template('learning_dashboard.html')

@app.route('/learning-stats')
def get_learning_stats():
    try:
        stats = ai_estimator.learning_system.get_accuracy_stats()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update-actual-hours', methods=['POST'])
def update_actual_hours():
    try:
        data = request.get_json()
        ticket_key = data.get('ticket_key')
        actual_hours = data.get('actual_hours')
        
        if not ticket_key or actual_hours is None:
            return jsonify({'error': 'Ticket key and actual hours are required'}), 400
        
        success = ai_estimator.learning_system.update_actual_hours(ticket_key, actual_hours)
        
        if success:
            return jsonify({'success': True, 'message': 'Actual hours updated successfully'})
        else:
            return jsonify({'error': 'Ticket not found in estimation history'}), 404
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/bulk-estimate', methods=['POST'])
def bulk_estimate():
    """Estimate multiple JIRA tickets at once"""
    try:
        data = request.get_json()
        ticket_numbers = data.get('ticket_numbers', [])
        use_ai = data.get('use_ai', False)
        uses_ai_tools = data.get('uses_ai_tools', False)
        
        if not ticket_numbers:
            return jsonify({'error': 'No ticket numbers provided'}), 400
        
        results = []
        
        for ticket_number in ticket_numbers:
            try:
                # Fetch JIRA data
                jira_data = jira_client.get_ticket_details(ticket_number.strip())
                description = f"{jira_data['summary']}. {jira_data['description']}"
                
                # Add AI tools flag
                jira_data['uses_ai_tools'] = uses_ai_tools
                
                # Estimate
                if use_ai:
                    estimation = ai_estimator.estimate_with_ai(description, jira_data)
                else:
                    estimation = estimator.estimate_project(description, ticket_number, jira_data)
                
                # Get release/fix version
                release = 'N/A'
                if jira_data.get('fix_versions'):
                    release = jira_data['fix_versions'][0]['name']
                
                results.append({
                    'ticket': ticket_number.strip(),
                    'release': release,
                    'complexity': estimation.get('complexity', 'Medium'),
                    'issue_type': jira_data.get('issue_type', 'Unknown'),
                    'total_hours': estimation.get('total_hours', 0),
                    'ai_confidence': estimation.get('ai_confidence', estimation.get('confidence', 0)),
                    'ai_estimate': 'Y' if use_ai else 'N',
                    'ai_tools': 'Y' if uses_ai_tools else 'N',
                    'phases': estimation.get('phases', {}),
                    'status': 'success'
                })
                
            except Exception as e:
                results.append({
                    'ticket': ticket_number.strip(),
                    'status': 'error',
                    'error': str(e)
                })
        
        return jsonify({'results': results})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export-excel', methods=['POST'])
def export_excel():
    """Export bulk estimates to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from flask import send_file
        
        data = request.get_json()
        results = data.get('results', [])
        
        if not results:
            return jsonify({'error': 'No results to export'}), 400
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Estimation Report"
        
        # Headers
        headers = [
            'Jira ticket', 'Release', 'Complexity', 'Issue Type', 'Total Estd. Hrs',
            'AI Confidence', 'AI Estimate', 'AI Tool Present', 'Requirement',
            'Design', 'Development', 'Testing', 'Deployment'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for row_idx, result in enumerate(results, 2):
            if result.get('status') == 'success':
                phases = result.get('phases', {})
                
                ws.cell(row=row_idx, column=1, value=result.get('ticket', ''))
                ws.cell(row=row_idx, column=2, value=result.get('release', 'N/A'))
                ws.cell(row=row_idx, column=3, value=result.get('complexity', 'Medium'))
                ws.cell(row=row_idx, column=4, value=result.get('issue_type', 'Unknown'))
                ws.cell(row=row_idx, column=5, value=round(result.get('total_hours', 0), 2))
                ws.cell(row=row_idx, column=6, value=f"{result.get('ai_confidence', 0)}%")
                ws.cell(row=row_idx, column=7, value=result.get('ai_estimate', 'N'))
                ws.cell(row=row_idx, column=8, value=result.get('ai_tools', 'N'))
                ws.cell(row=row_idx, column=9, value=round(phases.get('requirements', 0), 2))
                ws.cell(row=row_idx, column=10, value=round(phases.get('design', 0), 2))
                ws.cell(row=row_idx, column=11, value=round(phases.get('development', 0), 2))
                ws.cell(row=row_idx, column=12, value=round(phases.get('testing', 0), 2))
                ws.cell(row=row_idx, column=13, value=round(phases.get('deployment', 0), 2))
        
        # Adjust column widths
        for col in range(1, 14):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='estimation_report.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/test-ai', methods=['GET'])
def test_ai_connection():
    """Test Azure OpenAI FAB API connection"""
    try:
        # Test simple AI call
        test_result = ai_estimator.estimate_with_ai("Create a simple login form with username and password")
        
        return jsonify({
            'status': 'success',
            'ai_working': True,
            'estimation_method': test_result.get('estimation_method'),
            'test_hours': test_result.get('total_hours'),
            'message': 'AI connection working'
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'ai_working': False,
            'error': str(e),
            'message': 'AI connection failed'
        }), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)